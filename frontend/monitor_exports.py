"""
PDF ve Excel export fonksiyonları — monitor_app.py'den ayrı tutulur.
Streamlit magic yalnızca ana script'e uygulanır; buraya uygulanmaz.
"""
import io


def _fb_list(val) -> list:
    if isinstance(val, list):
        return val
    if isinstance(val, dict):
        try:
            return [val[str(i)] for i in range(len(val))]
        except KeyError:
            return list(val.values())
    return []


def gen_alarm_excel(d: dict) -> bytes | None:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        def _hdr(ws, cols):
            ws.append(cols)
            fill = PatternFill("solid", fgColor="1A2235")
            font = Font(bold=True, color="8AB8D8")
            algn = Alignment(horizontal="center")
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = algn
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = max(
                    len(str(col[0].value or "")), 10) + 2

        wb  = openpyxl.Workbook()
        dd  = d.get("daily_stats") or {}
        cs  = dd.get("cycles", {})
        als = dd.get("alarms", {})

        ws1 = wb.active
        ws1.title = "7 Gunluk Trend"
        _hdr(ws1, ["Tarih", "Cycle Sayisi", "Anomali Sayisi", "Alarm Sayisi", "Ort. Sure (s)"])
        for r in _fb_list(d.get("trend_7d")):
            ws1.append([r.get("date", ""), r.get("cycle_count", 0),
                        r.get("anomaly_count", 0), r.get("alarm_count", 0),
                        r.get("avg_duration", 0)])

        ws2 = wb.create_sheet("Alarm Gecmisi")
        _hdr(ws2, ["Alarm Kodu", "Ad", "Tekrar Sayisi"])
        for a in _fb_list(als.get("top_codes")):
            ws2.append([a.get("code", ""), a.get("name", ""), a.get("count", 0)])

        ws3 = wb.create_sheet("Gunluk Ozet")
        _hdr(ws3, ["Metrik", "Deger"])
        ws3.append(["Cycle Sayisi",  cs.get("count", 0)])
        ws3.append(["Alarm Sayisi",  als.get("count", 0)])
        ws3.append(["Anomali Orani", f'{cs.get("anomaly_rate", 0):.1f}%'])
        ws3.append(["Ort. Sure (s)", f'{cs.get("avg_duration", 0):.0f}'])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return None


def gen_alarm_pdf(d: dict, lang: str) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
    from datetime import date as _dt

    NL = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}
    _s = lambda t: t.translate(str.maketrans("ğĞüÜşŞıİöÖçÇ", "gGuUssiIoOcC")).encode("latin-1", errors="replace").decode("latin-1")

    if lang == "en":
        L = {"title": "Production Report", "cycles": "Cycles", "anom": "Anomaly Rate",
             "alarms": "Alarms", "dur": "Avg Duration (s)", "trend": "7-Day Trend",
             "th": "Most Frequent Alarms", "d": "Date", "cy": "Cycle", "an": "Anomaly",
             "al": "Alarm", "du": "Avg Dur.", "co": "Code", "na": "Name", "ct": "Count",
             "gen": "Generated", "daily": "Daily Summary", "anom_status": "Anomaly Status",
             "anom_ok": "No Anomaly Detected", "anom_warn": "Anomaly Detected",
             "anom_detail": "Detail", "total_cycles": "Total Cycles Analyzed",
             "alarm_hist": "Alarm History", "ah_time": "Time", "ah_code": "Code",
             "ah_sev": "Severity", "ah_type": "Type", "ah_name": "Name",
             "sev_high": "Critical", "sev_med": "Medium", "sev_low": "Low"}
    else:
        L = {"title": "Uretim Raporu", "cycles": "Cycle Sayisi", "anom": "Anomali Orani",
             "alarms": "Alarm Sayisi", "dur": "Ort. Sure (s)", "trend": "7 Gunluk Trend",
             "th": "En Cok Tekrar Eden Alarmlar", "d": "Tarih", "cy": "Cycle", "an": "Anomali",
             "al": "Alarm", "du": "Ort. Sure", "co": "Kod", "na": "Ad", "ct": "Sayi",
             "gen": "Olusturulma", "daily": "Gunluk Ozet", "anom_status": "Anomali Durumu",
             "anom_ok": "Anomali Tespit Edilmedi", "anom_warn": "Anomali Tespit Edildi",
             "anom_detail": "Detay", "total_cycles": "Analiz Edilen Cycle",
             "alarm_hist": "Alarm Gecmisi", "ah_time": "Zaman", "ah_code": "Kod",
             "ah_sev": "Onem", "ah_type": "Tur", "ah_name": "Ad",
             "sev_high": "Kritik", "sev_med": "Orta", "sev_low": "Dusuk"}

    dd = d.get("daily_stats") or {}
    cs = dd.get("cycles", {})
    als = dd.get("alarms", {})
    trend = _fb_list(d.get("trend_7d"))
    top = _fb_list(als.get("top_codes"))
    anom_latest = d.get("anomaly_latest") or {}
    alarm_hist = _fb_list(d.get("alarm_history"))

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=12)

    def _sec(title):
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(26, 34, 53)
        pdf.set_text_color(138, 184, 216)
        pdf.cell(0, 7, title, fill=True, **NL)
        pdf.ln(1)

    pdf.set_font("Helvetica", "B", 14)
    pdf.set_fill_color(26, 34, 53)
    pdf.set_text_color(138, 184, 216)
    pdf.cell(0, 11, L["title"], fill=True, align="C", **NL)
    pdf.ln(3)

    _sec(L["daily"])
    kpis = [(L["cycles"], str(cs.get("count", 0))),
            (L["anom"], f'{cs.get("anomaly_rate", 0):.1f}%'),
            (L["alarms"], str(als.get("count", 0))),
            (L["dur"], f'{cs.get("avg_duration", 0):.0f} s')]
    cw = pdf.epw / 4
    for k, v in kpis:
        x, y = pdf.get_x(), pdf.get_y()
        pdf.set_fill_color(240, 244, 250)
        pdf.rect(x, y, cw - 2, 15, "F")
        pdf.set_xy(x, y + 1)
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(cw - 2, 4, _s(k), align="C")
        pdf.set_xy(x, y + 6)
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(30, 80, 140)
        pdf.cell(cw - 2, 7, v, align="C")
        pdf.set_xy(x + cw, y)
    pdf.ln(18)

    _sec(L["anom_status"])
    if anom_latest:
        has_anom = anom_latest.get("has_anomaly", False)
        total_cy = anom_latest.get("total_cycles", 0)
        detail_lines = anom_latest.get("detail_lines", [])
        pdf.set_font("Helvetica", "B", 9)
        if has_anom:
            pdf.set_fill_color(180, 40, 40)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 7, f"! {_s(L['anom_warn'])}", fill=True, border=1, **NL)
            if detail_lines:
                pdf.set_font("Helvetica", "", 8)
                pdf.set_fill_color(255, 240, 240)
                pdf.set_text_color(40, 40, 40)
                for line in detail_lines:
                    pdf.cell(0, 5, f"  - {_s(str(line))}", fill=True, border=1, **NL)
        else:
            pdf.set_fill_color(20, 80, 40)
            pdf.set_text_color(200, 255, 200)
            pdf.cell(0, 7, f"OK  {_s(L['anom_ok'])}", fill=True, border=1, **NL)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(100, 100, 100)
        pdf.set_fill_color(245, 245, 245)
        pdf.cell(0, 5, f"  {_s(L['total_cycles'])}: {total_cy}", fill=True, **NL)
    else:
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 5, _s("Veri yok" if lang == "tr" else "No data"), **NL)
    pdf.ln(3)

    if trend:
        _sec(L["trend"])
        ww = [34, 24, 24, 24, 24]
        hs = [L["d"], L["cy"], L["an"], L["al"], L["du"]]
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(220, 230, 242)
        pdf.set_text_color(20, 20, 20)
        for c, w in zip(hs, ww):
            pdf.cell(w, 6, c, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for i, r in enumerate(trend):
            if i % 2 == 0:
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(242, 246, 252)
            pdf.set_text_color(40, 40, 40)
            for v, w in zip([r.get("date", ""), str(r.get("cycle_count", 0)),
                             str(r.get("anomaly_count", 0)), str(r.get("alarm_count", 0)),
                             f'{r.get("avg_duration", 0):.0f}'], ww):
                pdf.cell(w, 5, _s(str(v)), border=1, fill=True, align="C")
            pdf.ln()
        pdf.ln(3)

    if top:
        _sec(L["th"])
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(220, 230, 242)
        pdf.set_text_color(20, 20, 20)
        for c, w in zip([L["co"], L["na"], L["ct"]], [22, 130, 18]):
            pdf.cell(w, 6, c, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for i, a in enumerate(top):
            if i % 2 == 0:
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(242, 246, 252)
            pdf.set_text_color(40, 40, 40)
            for v, w in zip([a.get("code", ""), a.get("name", ""), str(a.get("count", 0))],
                            [22, 130, 18]):
                pdf.cell(w, 5, _s(str(v)), border=1, fill=True)
            pdf.ln()
        pdf.ln(3)

    if alarm_hist:
        _sec(L["alarm_hist"])
        sev_map = {"high": L["sev_high"], "medium": L["sev_med"], "low": L["sev_low"]}
        SEV_COLORS = {"high": (255, 235, 235), "medium": (255, 248, 220), "low": (248, 250, 252)}
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(26, 34, 53)
        pdf.set_text_color(138, 184, 216)
        nw = pdf.epw - 34 - 16 - 16 - 20
        for c, w in zip([L["ah_time"], L["ah_code"], L["ah_sev"], L["ah_type"], L["ah_name"]],
                        [34, 34, 16, 20, nw]):
            pdf.cell(w, 6, c, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for a in alarm_hist:
            sev = a.get("severity") or ""
            pdf.set_fill_color(*SEV_COLORS.get(sev, (248, 250, 252)))
            pdf.set_text_color(40, 40, 40)
            ts = _s(str(a.get("timestamp", ""))[:16].replace("T", " "))
            pdf.cell(34, 5, ts, border=1, fill=True)
            pdf.cell(16, 5, _s(str(a.get("code", ""))), border=1, fill=True, align="C")
            pdf.cell(16, 5, _s(sev_map.get(sev, sev)), border=1, fill=True, align="C")
            pdf.cell(20, 5, _s(str(a.get("type", ""))[:12]), border=1, fill=True, align="C")
            pdf.cell(nw, 5, _s(str(a.get("name", ""))[:50]), border=1, fill=True)
            pdf.ln()
        pdf.ln(3)

    pdf.set_y(-12)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 4, f'{L["gen"]}: {_dt.today().isoformat()} - Curing Press Monitor', align="C")

    return bytes(pdf.output())


def gen_energy_excel(d: dict) -> bytes | None:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        def _hdr(ws, cols):
            ws.append(cols)
            fill = PatternFill("solid", fgColor="1a3a5c")
            font = Font(bold=True, color="a8c8e8")
            algn = Alignment(horizontal="center")
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = algn
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = max(
                    len(str(col[0].value or "")), 12) + 2

        wb = openpyxl.Workbook()
        tod = d.get("energy_today") or {}

        ws1 = wb.active
        ws1.title = "Saatlik Ozet"
        _hdr(ws1, ["Saat", "Ort. kW", "Ort. m3/h"])
        for r in _fb_list(d.get("energy_trend")):
            ws1.append([r.get("hour", ""), r.get("kw_mean", 0), r.get("air_flow_mean", 0)])
        ws1.append([])
        ws1.append(["Gunluk Toplam kWh", tod.get("kwh_total", 0)])
        ws1.append(["Gunluk Toplam m3",  tod.get("nm3_total", 0)])
        ws1.append(["kWh/Lastik",         tod.get("kwh_per_tire", 0)])
        ws1.append(["m3/Lastik",           tod.get("nm3_per_tire", 0)])

        ws2 = wb.create_sheet("Vardiya Ozeti")
        _hdr(ws2, ["Vardiya", "kWh", "m3", "Ort. kW", "Calisma %"])
        for r in _fb_list(d.get("shift_summary")):
            ws2.append([r.get("shift", ""), r.get("kwh_total", 0), r.get("nm3_total", 0),
                        r.get("kw_mean", 0), r.get("running_pct", 0)])

        ws3 = wb.create_sheet("Cycle Bazli")
        _hdr(ws3, ["Cycle ID", "Baslangic", "kWh/Cycle", "m3/Cycle",
                   "Ort. kW", "Maks. kW", "Ort. m3/h"])
        for r in _fb_list(d.get("cycle_energy")):
            ws3.append([r.get("cycle_id", ""), r.get("start_time", ""),
                        r.get("kwh_cycle", 0), r.get("nm3_cycle", 0),
                        r.get("kw_mean", 0), r.get("kw_max", 0),
                        r.get("air_flow_mean", 0)])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return None


def gen_energy_pdf(d: dict, lang: str) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
    from datetime import date as _dt

    NL = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}
    _s = lambda t: t.translate(str.maketrans("ğĞüÜşŞıİöÖçÇ", "gGuUssiIoOcC")).encode("latin-1", errors="replace").decode("latin-1")

    if lang == "en":
        L = {"title": "Energy Consumption Report", "kpi": "Daily Summary", "kwh": "Total kWh",
             "nm3": "Total m3", "kwht": "kWh/Tire", "nm3t": "m3/Tire", "shift": "Shift Summary",
             "cycle": "Cycle Energy (Last 20)", "hs": "Shift", "hk": "kWh", "hn": "m3",
             "hkw": "Avg kW", "hr": "Running %", "hci": "Cycle ID", "hst": "Start",
             "hck": "kWh/Cycle", "hcn": "m3/Cycle", "hmk": "Max kW", "ham": "Avg m3/h",
             "hourly": "Hourly Trend (Last 24h)", "hhr": "Hour", "hkwh": "Avg kW",
             "ham2": "Avg m3/h", "gen": "Generated"}
    else:
        L = {"title": "Enerji Tuketim Raporu", "kpi": "Gunluk Ozet", "kwh": "Toplam kWh",
             "nm3": "Toplam m3", "kwht": "kWh/Lastik", "nm3t": "m3/Lastik",
             "shift": "Vardiya Ozeti", "cycle": "Cycle Bazli Enerji (Son 20)",
             "hs": "Vardiya", "hk": "kWh", "hn": "m3", "hkw": "Ort. kW", "hr": "Calisma %",
             "hci": "Cycle ID", "hst": "Baslangic", "hck": "kWh/Cycle", "hcn": "m3/Cycle",
             "hmk": "Maks. kW", "ham": "Ort. m3/h", "hourly": "Saatlik Trend (Son 24 Saat)",
             "hhr": "Saat", "hkwh": "Ort. kW", "ham2": "Ort. m3/h", "gen": "Olusturulma"}

    tod = d.get("energy_today") or {}
    sv  = _fb_list(d.get("shift_summary"))
    ec  = _fb_list(d.get("cycle_energy"))
    et  = _fb_list(d.get("energy_trend"))

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_fill_color(26, 58, 92)
    pdf.set_text_color(168, 200, 232)
    pdf.cell(0, 12, L["title"], fill=True, align="C", **NL)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(26, 34, 53)
    pdf.set_text_color(138, 184, 216)
    pdf.cell(0, 9, L["kpi"], fill=True, **NL)
    pdf.set_font("Helvetica", "B", 10)
    for k, v in [(L["kwh"],  f'{tod.get("kwh_total",  0):.2f} kWh'),
                 (L["nm3"],  f'{tod.get("nm3_total",  0):.3f} m3'),
                 (L["kwht"], f'{tod.get("kwh_per_tire",0):.3f} kWh/tire'),
                 (L["nm3t"], f'{tod.get("nm3_per_tire",0):.4f} m3/tire')]:
        pdf.set_fill_color(26, 34, 53)
        pdf.set_text_color(138, 184, 216)
        pdf.cell(90, 8, k, fill=True, border=1)
        pdf.set_fill_color(240, 244, 248)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(60, 8, v, fill=True, border=1, **NL)
    pdf.ln(5)

    if et:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(26, 34, 53)
        pdf.set_text_color(138, 184, 216)
        pdf.cell(0, 9, L["hourly"], fill=True, **NL)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(220, 230, 242)
        pdf.set_text_color(20, 20, 20)
        for c, w in zip([L["hhr"], L["hkwh"], L["ham2"]], [50, 65, 65]):
            pdf.cell(w, 7, c, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 9)
        for i, r in enumerate(et):
            if i % 2 == 0:
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(240, 248, 255)
            pdf.set_text_color(40, 40, 40)
            for v, w in zip([str(r.get("hour", "")),
                             f'{r.get("kw_mean", 0):.2f}',
                             f'{r.get("air_flow_mean", 0):.2f}'], [50, 65, 65]):
                pdf.cell(w, 6, _s(str(v)), border=1, fill=True, align="C")
            pdf.ln()
        pdf.ln(4)

    if sv:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(26, 34, 53)
        pdf.set_text_color(138, 184, 216)
        pdf.cell(0, 9, L["shift"], fill=True, **NL)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(220, 230, 242)
        pdf.set_text_color(20, 20, 20)
        for c, w in zip([L["hs"], L["hk"], L["hn"], L["hkw"], L["hr"]], [30, 32, 32, 32, 32]):
            pdf.cell(w, 7, c, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 9)
        for i, r in enumerate(sv):
            if i % 2 == 0:
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(240, 248, 255)
            pdf.set_text_color(40, 40, 40)
            for v, w in zip([r.get("shift", ""),
                             f'{r.get("kwh_total", 0):.3f}',
                             f'{r.get("nm3_total", 0):.3f}',
                             f'{r.get("kw_mean", 0):.1f}',
                             f'{r.get("running_pct", 0):.1f}%'], [30, 32, 32, 32, 32]):
                pdf.cell(w, 6, _s(str(v)), border=1, fill=True, align="C")
            pdf.ln()
        pdf.ln(4)

    if ec:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(26, 34, 53)
        pdf.set_text_color(138, 184, 216)
        pdf.cell(0, 9, L["cycle"], fill=True, **NL)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(220, 230, 242)
        pdf.set_text_color(20, 20, 20)
        for c, w in zip([L["hci"], L["hst"], L["hck"], L["hcn"],
                         L["hkw"], L["hmk"], L["ham"]], [18, 42, 28, 28, 28, 28, 18]):
            pdf.cell(w, 7, c, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for i, r in enumerate(ec[-20:]):
            if i % 2 == 0:
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(240, 248, 255)
            pdf.set_text_color(40, 40, 40)
            for v, w in zip([str(r.get("cycle_id", "")),
                             str(r.get("start_time", ""))[:16],
                             f'{r.get("kwh_cycle", 0):.3f}',
                             f'{r.get("nm3_cycle", 0):.4f}',
                             f'{r.get("kw_mean", 0):.1f}',
                             f'{r.get("kw_max", 0):.1f}',
                             f'{r.get("air_flow_mean", 0):.1f}'],
                            [18, 42, 28, 28, 28, 28, 18]):
                pdf.cell(w, 6, _s(str(v)), border=1, fill=True, align="C")
            pdf.ln()

    pdf.set_y(-15)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, f'{L["gen"]}: {_dt.today().isoformat()} - Curing Press Monitor', align="C")

    return bytes(pdf.output())
